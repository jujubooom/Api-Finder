#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
输出管理模块 (Output Manager Module)
负责处理Api-Finder的所有输出功能，包括终端输出和文件输出
"""

import os
import json
import time
from datetime import datetime
from urllib.parse import urlparse
from rich.console import Console
from rich.text import Text
from rich.panel import Panel
from rich.table import Table
from rich.progress import Progress, SpinnerColumn, TextColumn, BarColumn, TimeElapsedColumn
from rich.rule import Rule
from .i18n import i18n


class OutputManager:
    """
    使用Rich库的输出管理器类，提供美观的终端输出和多种文件输出格式
    
    Attributes:
        silent_mode (bool): 静默模式，只输出发现的API端点
        verbose_mode (bool): 详细输出模式
        output_file (str): 输出文件路径
        results (list): 结果列表
        stats (dict): 统计信息
        console (Console): Rich console对象
        results_table (Table): 结果表格
    """
    
    def __init__(self, silent_mode, verbose_mode=False, output_file=None):
        """
        初始化输出管理器
        
        Args:
            silent_mode (bool): 静默模式
            verbose_mode (bool): 详细输出模式
            output_file (str): 输出文件路径
        """
        self.silent_mode = silent_mode
        self.verbose_mode = verbose_mode
        self.output_file = output_file
        self.console = Console()
        self.results = []
        self.stats = {
            "total_urls": 0,
            "successful_requests": 0,
            "failed_requests": 0,
            "api_endpoints": 0,
            "start_time": datetime.now()
        }
        self.results_table = Table(title="🔍 Discovered API Endpoints", border_style="green")
        self.results_table.add_column("📍 URL", style="cyan", no_wrap=False)
        self.results_table.add_column("📄 Source", style="yellow", max_width=30)
        self.results_table.add_column("⏰ Time", style="dim", max_width=10)
    
    def print_info(self, text):
        """打印信息"""
        if not self.silent_mode:
            self.console.print(text)
    
    def print_verbose(self, text):
        """打印详细信息"""
        if self.verbose_mode and not self.silent_mode:
            self.console.print(f"[dim][DEBUG][/dim] {text}")
    
    def print_url(self, url, source=""):
        """打印发现的URL"""
        if self.silent_mode:
            # 静默模式使用Rich的print而不是普通print
            self.console.print(url, highlight=False)
        else:
            # 添加到结果表格
            source_display = source.split('/')[-1] if source else "unknown"
            time_display = datetime.now().strftime("%H:%M:%S")
            self.results_table.add_row(url, source_display, time_display)
            
            if source:
                self.console.print(f"[green bold]✓[/green bold] [blue]{url}[/blue] [dim](from: {source_display})[/dim]")
            else:
                self.console.print(f"[green bold]✓[/green bold] [blue]{url}[/blue]")
        
        # 保存结果
        self.results.append({
            "url": url,
            "source": source,
            "timestamp": datetime.now().isoformat()
        })
        self.stats["api_endpoints"] += 1
    
    def print_error(self, text):
        """打印错误信息"""
        if not self.silent_mode:
            self.console.print(f"[red bold]✗[/red bold] {text}")
    
    def print_warning(self, text):
        """打印警告信息"""
        if not self.silent_mode:
            self.console.print(f"[yellow bold]⚠[/yellow bold] {text}")
    
    def print_success(self, text):
        """打印成功信息"""
        if not self.silent_mode:
            self.console.print(f"[green bold]✓[/green bold] {text}")

    def print_title(self, url, title):
        """打印成功请求的页面标题"""
        if not self.silent_mode:
            text = Text()
            text.append("📄 ", style="green")
            text.append(f"{title}", style="yellow")
            text.append(" → ", style="dim")
            text.append(f"{url}", style="cyan dim")
            self.console.print(text)

    def print_proxy_mode(self, proxies):
        """输出使用的代理模式"""
        if not self.silent_mode:
            if proxies:
                proxy_table = Table(title="🌐 Proxy Configuration", border_style="blue")
                proxy_table.add_column("Type", style="cyan")
                proxy_table.add_column("Address", style="green")
                
                if isinstance(proxies, list):
                    for proxy in proxies:
                        proxy_table.add_row("SOCKS5", proxy)
                elif isinstance(proxies, dict):
                    for protocol, proxy in proxies.items():
                        proxy_table.add_row(protocol.upper(), proxy)
                
                self.console.print(proxy_table)
            else:
                self.console.print("[yellow]💻 Direct connection (no proxy)[/yellow]")
            self.console.print(Rule(style="dim"))

    def print_stats(self):
        """打印统计信息"""
        if not self.silent_mode:
            # 计算扫描时间
            scan_duration = datetime.now() - self.stats["start_time"]
            duration_str = f"{scan_duration.total_seconds():.1f}s"
            
            # 创建统计表格
            stats_table = Table(title="📊 Scan Statistics", border_style="cyan")
            stats_table.add_column("Item", style="yellow bold")
            stats_table.add_column("Value", style="green bold", justify="right")
            
            stats_table.add_row("🎯 Total URLs", str(self.stats['total_urls']))
            stats_table.add_row("✅ Successful Requests", str(self.stats['successful_requests']))
            stats_table.add_row("❌ Failed Requests", str(self.stats['failed_requests']))
            stats_table.add_row("🔍 API Endpoints Found", str(self.stats['api_endpoints']))
            stats_table.add_row("⏱️ Scan Duration", duration_str)
            
            # 计算成功率
            total_requests = self.stats['successful_requests'] + self.stats['failed_requests']
            if total_requests > 0:
                success_rate = (self.stats['successful_requests'] / total_requests) * 100
                stats_table.add_row("📈 Success Rate", f"{success_rate:.1f}%")
            
            self.console.print(Rule(style="dim"))
            self.console.print(stats_table)
            
            # 如果找到了API端点，显示结果表格
            if self.stats['api_endpoints'] > 0 and not self.silent_mode:
                self.console.print(Rule(style="dim"))
                self.console.print(self.results_table)
    
    def create_progress(self, total_tasks=None):
        """创建进度条"""
        if self.silent_mode:
            return None
        
        return Progress(
            SpinnerColumn(),
            TextColumn("[progress.description]{task.description}"),
            BarColumn(),
            TextColumn("[progress.percentage]{task.percentage:>3.0f}%"),
            TimeElapsedColumn(),
            console=self.console,
            expand=True
        )


class FileOutputManager:
    """
    文件输出管理器类
    负责处理各种文件格式的输出
    """
    
    def __init__(self, output_manager):
        """
        初始化文件输出管理器
        
        Args:
            output_manager (OutputManager): 输出管理器实例
        """
        self.output_manager = output_manager
        self.console = output_manager.console
    
    def save_results(self, target_url, config_args):
        """
        保存扫描结果到文件
        
        Args:
            target_url (str): 目标URL
            config_args: 配置参数对象
        """
        if not self.output_manager.output_file:
            return
        
        try:
            # 创建输出目录（如果不存在）
            output_dir = os.path.dirname(self.output_manager.output_file)
            if output_dir and not os.path.exists(output_dir):
                os.makedirs(output_dir)
            
            file_ext = os.path.splitext(self.output_manager.output_file)[1].lower()
            
            # 数据去重和排序
            unique_results = self._deduplicate_results()
            sorted_results = self._sort_results(unique_results)
            
            # 根据文件扩展名选择输出格式
            if file_ext == '.json':
                self._save_as_json(sorted_results, target_url, config_args)
            elif file_ext == '.txt':
                self._save_as_txt(sorted_results, target_url)
            elif file_ext == '.csv':
                self._save_as_csv(sorted_results)
            elif file_ext == '.html':
                self._save_as_html(sorted_results, target_url)
            elif file_ext == '.xml':
                self._save_as_xml(sorted_results, target_url)
            elif file_ext == '.xlsx':
                self._save_as_excel(sorted_results, target_url)
            elif file_ext == '.md':
                self._save_as_markdown(sorted_results, target_url)
            else:
                # 默认保存为JSON格式
                self.output_manager.output_file = self.output_manager.output_file.rsplit('.', 1)[0] + '.json'
                self._save_as_json(sorted_results, target_url, config_args)
                self.output_manager.print_warning(f"不支持的文件格式，已保存为JSON格式")
            
            # 输出文件信息
            file_size = os.path.getsize(self.output_manager.output_file)
            file_size_str = self._format_file_size(file_size)
            
            if not self.output_manager.silent_mode:
                self.console.print(f"\n[green bold]💾 Results saved to:[/green bold] [blue]{self.output_manager.output_file}[/blue]")
                self.console.print(f"[dim]📁 File size: {file_size_str} | URLs: {len(sorted_results)} | Unique: {len(unique_results)} total[/dim]")
                
        except Exception as e:
            self.output_manager.print_error(f"Save failed: {str(e)}")
    
    def _deduplicate_results(self):
        """去重结果"""
        seen_urls = set()
        unique_results = []
        
        for result in self.output_manager.results:
            url = result['url']
            if url not in seen_urls:
                seen_urls.add(url)
                unique_results.append(result)
        
        return unique_results
    
    def _sort_results(self, results):
        """排序结果 - 按URL字母顺序"""
        return sorted(results, key=lambda x: x['url'])
    
    def _format_file_size(self, size_bytes):
        """格式化文件大小"""
        if size_bytes == 0:
            return "0 B"
        
        size_names = ["B", "KB", "MB", "GB"]
        i = 0
        while size_bytes >= 1024 and i < len(size_names) - 1:
            size_bytes /= 1024.0
            i += 1
        
        return f"{size_bytes:.1f} {size_names[i]}"
    
    def _save_as_json(self, results, target_url, config_args):
        """保存为JSON格式"""
        scan_duration = datetime.now() - self.output_manager.stats["start_time"]
        
        output_data = {
            "metadata": {
                "version": "0.3.1",
                "tool": "Api-Finder",
                "scan_time": datetime.now().isoformat(),
                "target_url": target_url,
                "scan_duration_seconds": scan_duration.total_seconds(),
                "proxy_used": getattr(config_args, 'proxy', None) if config_args else "Direct",
                "total_results": len(results),
                "unique_results": len(self._deduplicate_results())
            },
            "statistics": {
                **self.output_manager.stats,
                "start_time": self.output_manager.stats["start_time"].isoformat(),
                "success_rate": round((self.output_manager.stats["successful_requests"] / max(1, self.output_manager.stats["successful_requests"] + self.output_manager.stats["failed_requests"])) * 100, 2)
            },
            "results": results,
            "configuration": {
                "timeout": getattr(config_args, 'timeout', 10) if config_args else 10,
                "delay": getattr(config_args, 'delay', 0.5) if config_args else 0.5,
                "verbose": getattr(config_args, 'verbose', False) if config_args else False,
                "silent": getattr(config_args, 'silent', False) if config_args else False,
                "random_ua": getattr(config_args, 'random', False) if config_args else False
            }
        }
        
        with open(self.output_manager.output_file, 'w', encoding='utf-8') as f:
            json.dump(output_data, f, ensure_ascii=False, indent=2)
    
    def _save_as_txt(self, results, target_url):
        """保存为TXT格式"""
        with open(self.output_manager.output_file, 'w', encoding='utf-8') as f:
            # 写入文件头
            f.write("=" * 60 + "\n")
            f.write(f"{i18n.get('output_header')}\n")
            f.write("=" * 60 + "\n")
            f.write(f"{i18n.get('output_target')}: {target_url}\n")
            f.write(f"{i18n.get('output_scan_time')}: {datetime.now().isoformat()}\n")
            f.write(f"扫描用时: {(datetime.now() - self.output_manager.stats['start_time']).total_seconds():.1f}秒\n")
            f.write(f"{i18n.get('output_endpoints_found')}: {len(results)}\n")
            f.write(f"成功请求: {self.output_manager.stats['successful_requests']}\n")
            f.write(f"失败请求: {self.output_manager.stats['failed_requests']}\n")
            f.write("-" * 60 + "\n\n")
            
            # 按来源分组输出
            sources = {}
            for result in results:
                source = result['source'] if result['source'] else 'Unknown'
                if source not in sources:
                    sources[source] = []
                sources[source].append(result)
            
            for source, source_results in sources.items():
                f.write(f"📁 来源: {source}\n")
                f.write("-" * 30 + "\n")
                for result in source_results:
                    f.write(f"{result['url']}\n")
                f.write("\n")
    
    def _save_as_csv(self, results):
        """保存为CSV格式"""
        import csv
        with open(self.output_manager.output_file, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            # 写入头部
            writer.writerow(['URL', 'Source', 'Timestamp', 'Source_Type', 'Domain'])
            
            for result in results:
                url = result['url']
                source = result['source'] if result['source'] else 'Unknown'
                timestamp = result['timestamp']
                
                # 分析URL类型
                url_type = self._analyze_url_type(url)
                
                # 提取域名
                try:
                    domain = urlparse(url).netloc
                except:
                    domain = 'Unknown'
                
                writer.writerow([url, source, timestamp, url_type, domain])
    
    def _save_as_html(self, results, target_url):
        """保存为HTML格式"""
        html_content = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>API Finder - 扫描结果</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; background-color: #f5f5f5; }}
        .container {{ max-width: 1200px; margin: 0 auto; background: white; padding: 20px; border-radius: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }}
        h1 {{ color: #333; text-align: center; }}
        .stats {{ display: flex; justify-content: space-around; margin: 20px 0; }}
        .stat {{ text-align: center; padding: 10px; background: #e8f4f8; border-radius: 4px; }}
        .stat-value {{ font-size: 24px; font-weight: bold; color: #2196F3; }}
        .stat-label {{ font-size: 14px; color: #666; }}
        table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
        th, td {{ padding: 12px; text-align: left; border-bottom: 1px solid #ddd; }}
        th {{ background-color: #f8f9fa; font-weight: bold; }}
        .url-link {{ color: #2196F3; text-decoration: none; }}
        .url-link:hover {{ text-decoration: underline; }}
        .source {{ color: #666; font-size: 12px; }}
        .timestamp {{ color: #888; font-size: 11px; }}
        .filter-box {{ margin: 20px 0; }}
        .filter-box input {{ padding: 8px; border: 1px solid #ddd; border-radius: 4px; width: 300px; }}
    </style>
    <script>
        function filterResults() {{
            const input = document.getElementById('filterInput');
            const filter = input.value.toLowerCase();
            const table = document.getElementById('resultsTable');
            const rows = table.getElementsByTagName('tr');
            
            for (let i = 1; i < rows.length; i++) {{
                const url = rows[i].getElementsByTagName('td')[0].textContent.toLowerCase();
                if (url.indexOf(filter) > -1) {{
                    rows[i].style.display = '';
                }} else {{
                    rows[i].style.display = 'none';
                }}
            }}
        }}
    </script>
</head>
<body>
    <div class="container">
        <h1>🔍 API Finder 扫描结果</h1>
        
        <div class="stats">
            <div class="stat">
                <div class="stat-value">{len(results)}</div>
                <div class="stat-label">发现的URL</div>
            </div>
            <div class="stat">
                <div class="stat-value">{self.output_manager.stats['successful_requests']}</div>
                <div class="stat-label">成功请求</div>
            </div>
            <div class="stat">
                <div class="stat-value">{self.output_manager.stats['failed_requests']}</div>
                <div class="stat-label">失败请求</div>
            </div>
            <div class="stat">
                <div class="stat-value">{(datetime.now() - self.output_manager.stats['start_time']).total_seconds():.1f}s</div>
                <div class="stat-label">扫描用时</div>
            </div>
        </div>
        
        <p><strong>目标URL:</strong> {target_url}</p>
        <p><strong>扫描时间:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
        
        <div class="filter-box">
            <input type="text" id="filterInput" placeholder="过滤URL..." onkeyup="filterResults()">
        </div>
        
        <table id="resultsTable">
            <thead>
                <tr>
                    <th>URL</th>
                    <th>来源</th>
                    <th>类型</th>
                    <th>时间</th>
                </tr>
            </thead>
            <tbody>
"""
        
        for result in results:
            url = result['url']
            source = result['source'] if result['source'] else 'Unknown'
            timestamp = result['timestamp']
            url_type = self._analyze_url_type(url)
            
            # 格式化时间
            try:
                time_obj = datetime.fromisoformat(timestamp)
                formatted_time = time_obj.strftime('%H:%M:%S')
            except:
                formatted_time = timestamp
            
            html_content += f"""
                <tr>
                    <td><a href="{url}" class="url-link" target="_blank">{url}</a></td>
                    <td><span class="source">{source.split('/')[-1] if source else 'Unknown'}</span></td>
                    <td>{url_type}</td>
                    <td><span class="timestamp">{formatted_time}</span></td>
                </tr>
"""
        
        html_content += """
            </tbody>
        </table>
    </div>
</body>
</html>
"""
        
        with open(self.output_manager.output_file, 'w', encoding='utf-8') as f:
            f.write(html_content)
    
    def _save_as_xml(self, results, target_url):
        """保存为XML格式"""
        xml_content = f"""<?xml version="1.0" encoding="UTF-8"?>
<api_scan_results>
    <metadata>
        <tool>Api-Finder</tool>
        <version>0.3.1</version>
        <scan_time>{datetime.now().isoformat()}</scan_time>
        <target_url>{target_url}</target_url>
        <total_results>{len(results)}</total_results>
    </metadata>
    <statistics>
        <successful_requests>{self.output_manager.stats['successful_requests']}</successful_requests>
        <failed_requests>{self.output_manager.stats['failed_requests']}</failed_requests>
        <api_endpoints>{self.output_manager.stats['api_endpoints']}</api_endpoints>
        <scan_duration>{(datetime.now() - self.output_manager.stats['start_time']).total_seconds():.1f}</scan_duration>
    </statistics>
    <results>
"""
        
        for result in results:
            url = result['url']
            source = result['source'] if result['source'] else 'Unknown'
            timestamp = result['timestamp']
            url_type = self._analyze_url_type(url)
            
            xml_content += f"""
        <result>
            <url><![CDATA[{url}]]></url>
            <source><![CDATA[{source}]]></source>
            <type>{url_type}</type>
            <timestamp>{timestamp}</timestamp>
        </result>
"""
        
        xml_content += """
    </results>
</api_scan_results>
"""
        
        with open(self.output_manager.output_file, 'w', encoding='utf-8') as f:
            f.write(xml_content)
    
    def _save_as_excel(self, results, target_url):
        """保存为Excel格式"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "API扫描结果"
            
            # 设置标题样式
            title_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
            
            # 写入头部信息
            ws['A1'] = 'API Finder 扫描结果'
            ws['A1'].font = Font(bold=True, size=16)
            ws['A2'] = f'目标URL: {target_url}'
            ws['A3'] = f'扫描时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
            ws['A4'] = f'发现URL数量: {len(results)}'
            
            # 设置表格头部
            headers = ['URL', '来源', '类型', '域名', '时间戳']
            for i, header in enumerate(headers, 1):
                cell = ws.cell(row=6, column=i, value=header)
                cell.font = title_font
                cell.fill = header_fill
            
            # 写入数据
            for row, result in enumerate(results, 7):
                ws.cell(row=row, column=1, value=result['url'])
                ws.cell(row=row, column=2, value=result['source'] if result['source'] else 'Unknown')
                ws.cell(row=row, column=3, value=self._analyze_url_type(result['url']))
                try:
                    domain = urlparse(result['url']).netloc
                except:
                    domain = 'Unknown'
                ws.cell(row=row, column=4, value=domain)
                ws.cell(row=row, column=5, value=result['timestamp'])
            
            # 自动调整列宽
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(self.output_manager.output_file)
            
        except ImportError:
            self.output_manager.print_error("需要安装openpyxl库才能保存Excel格式: pip install openpyxl")
            # 回退到CSV格式
            self.output_manager.output_file = self.output_manager.output_file.rsplit('.', 1)[0] + '.csv'
            self._save_as_csv(results)
    
    def _save_as_markdown(self, results, target_url):
        """保存为Markdown格式"""
        md_content = f"""# 🔍 API Finder 扫描结果

## 📊 扫描信息

- **目标URL**: {target_url}
- **扫描时间**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
- **发现URL数量**: {len(results)}
- **成功请求**: {self.output_manager.stats['successful_requests']}
- **失败请求**: {self.output_manager.stats['failed_requests']}
- **扫描用时**: {(datetime.now() - self.output_manager.stats['start_time']).total_seconds():.1f}秒

## 🎯 发现的API端点

| URL | 来源 | 类型 | 时间 |
|-----|------|------|------|
"""
        
        for result in results:
            url = result['url']
            source = result['source'] if result['source'] else 'Unknown'
            source_display = source.split('/')[-1] if source else 'Unknown'
            url_type = self._analyze_url_type(url)
            
            try:
                time_obj = datetime.fromisoformat(result['timestamp'])
                formatted_time = time_obj.strftime('%H:%M:%S')
            except:
                formatted_time = result['timestamp']
            
            md_content += f"| {url} | {source_display} | {url_type} | {formatted_time} |\n"
        
        md_content += f"""

## 📈 统计信息

- 总URL数量: {len(results)}
- 唯一URL数量: {len(self._deduplicate_results())}
- 成功率: {round((self.output_manager.stats['successful_requests'] / max(1, self.output_manager.stats['successful_requests'] + self.output_manager.stats['failed_requests'])) * 100, 2)}%

---
*生成工具: Api-Finder v0.3.1*
"""
        
        with open(self.output_manager.output_file, 'w', encoding='utf-8') as f:
            f.write(md_content)
    
    def _analyze_url_type(self, url):
        """分析URL类型"""
        url_lower = url.lower()
        
        if any(keyword in url_lower for keyword in ['api', 'rest', 'graphql']):
            return 'API'
        elif url_lower.endswith('.json'):
            return 'JSON'
        elif url_lower.endswith('.xml'):
            return 'XML'
        elif any(keyword in url_lower for keyword in ['ajax', 'xhr']):
            return 'AJAX'
        elif any(keyword in url_lower for keyword in ['.php', '.jsp', '.asp']):
            return 'Dynamic'
        elif any(keyword in url_lower for keyword in ['.js', '.css', '.html']):
            return 'Static'
        else:
            return 'Other' 