import os
import json
from datetime import datetime
from urllib.parse import urlparse
from .i18n import i18n

class FileOutputManager:
    """
    文件输出管理器类
    负责处理各种文件格式的输出
    """

    def __init__(self, output_manager):
        """
        初始化文件输出管理器

        Args:
            output_manager (OutputManager): 输出管理器实例
        """
        self.output_manager = output_manager
        self.console = output_manager.console

    def save_results(self, target_url, config_args):
        """
        保存扫描结果到文件

        Args:
            target_url (str): 目标URL
            config_args: 配置参数对象
        """
        if not self.output_manager.output_file:
            return

        try:
            # 创建输出目录（如果不存在）
            output_dir = os.path.dirname(self.output_manager.output_file)
            if output_dir and not os.path.exists(output_dir):
                os.makedirs(output_dir)

            file_ext = os.path.splitext(self.output_manager.output_file)[1].lower()

            # 数据去重和排序
            unique_results = self._deduplicate_results()
            sorted_results = self._sort_results(unique_results)

            # 根据文件扩展名选择输出格式
            if file_ext == '.json':
                self._save_as_json(sorted_results, target_url, config_args)
            elif file_ext == '.txt':
                self._save_as_txt(sorted_results, target_url)
            elif file_ext == '.csv':
                self._save_as_csv(sorted_results)
            elif file_ext == '.html':
                self._save_as_html(sorted_results, target_url)
            elif file_ext == '.xml':
                self._save_as_xml(sorted_results, target_url)
            elif file_ext == '.xlsx':
                self._save_as_excel(sorted_results, target_url)
            elif file_ext == '.md':
                self._save_as_markdown(sorted_results, target_url)
            else:
                # 默认保存为JSON格式
                self.output_manager.output_file = self.output_manager.output_file.rsplit('.', 1)[0] + '.json'
                self._save_as_json(sorted_results, target_url, config_args)
                self.output_manager.print_warning(f"不支持的文件格式，已保存为JSON格式")

            # 输出文件信息
            file_size = os.path.getsize(self.output_manager.output_file)
            file_size_str = self._format_file_size(file_size)

            if not self.output_manager.silent_mode:
                self.console.print(
                    f"\n[green bold]💾All Results saved to:[/green bold] [blue]{self.output_manager.output_file}[/blue]")
                self.console.print(
                    f"[dim]📁 File size: {file_size_str} | URLs: {len(sorted_results)} | Unique: {len(unique_results)} total[/dim]")

        except Exception as e:
            self.output_manager.print_error(f"Save failed: {str(e)}")

    def _deduplicate_results(self):
        """去重结果"""
        seen_urls = set()
        unique_results = []

        for result in self.output_manager.results:
            url = result['url']
            if url not in seen_urls:
                seen_urls.add(url)
                unique_results.append(result)

        return unique_results

    def _sort_results(self, results):
        """排序结果 - 按URL字母顺序"""
        return sorted(results, key=lambda x: x['url'])

    def _format_file_size(self, size_bytes):
        """格式化文件大小"""
        if size_bytes == 0:
            return "0 B"

        size_names = ["B", "KB", "MB", "GB"]
        i = 0
        while size_bytes >= 1024 and i < len(size_names) - 1:
            size_bytes /= 1024.0
            i += 1

        return f"{size_bytes:.1f} {size_names[i]}"

    def _save_as_json(self, results, target_url, config_args):
        """保存为JSON格式"""
        scan_duration = datetime.now() - self.output_manager.stats["start_time"]

        output_data = {
            "metadata": {
                "version": "1,0",
                "tool": "Api-Finder",
                "scan_time": datetime.now().isoformat(),
                "target_url": target_url,
                "scan_duration_seconds": scan_duration.total_seconds(),
                "proxy_used": getattr(config_args, 'proxy', None) if config_args else "Direct",
                "total_results": len(results),
                "unique_results": len(self._deduplicate_results())
            },
            "statistics": {
                **self.output_manager.stats,
                "start_time": self.output_manager.stats["start_time"].isoformat(),
                "success_rate": round((self.output_manager.stats["successful_requests"] / max(1,
                                                                                              self.output_manager.stats[
                                                                                                  "successful_requests"] +
                                                                                              self.output_manager.stats[
                                                                                                  "failed_requests"])) * 100,
                                      2)
            },
            "results": results,
            "configuration": {
                "timeout": getattr(config_args, 'timeout', 10) if config_args else 10,
                "delay": getattr(config_args, 'delay', 0.5) if config_args else 0.5,
                "verbose": getattr(config_args, 'verbose', False) if config_args else False,
                "silent": getattr(config_args, 'silent', False) if config_args else False,
                "random_ua": getattr(config_args, 'random', False) if config_args else False
            }
        }

        with open(self.output_manager.output_file, 'w', encoding='utf-8') as f:
            json.dump(output_data, f, ensure_ascii=False, indent=2)


    # TXT
    def _save_as_txt(self, results, target_url):
        """保存为TXT格式"""
        with open(self.output_manager.output_file, 'w', encoding='utf-8') as f:
            # 写入文件头
            f.write("=" * 60 + "\n")
            f.write(f"{i18n.get('output_header')}\n")
            f.write("=" * 60 + "\n")
            f.write(f"{i18n.get('output_target')}: {target_url}\n")
            f.write(f"{i18n.get('output_scan_time')}: {datetime.now().isoformat()}\n")
            f.write(f"扫描用时: {(datetime.now() - self.output_manager.stats['start_time']).total_seconds():.1f}秒\n")
            f.write(f"{i18n.get('output_endpoints_found')}: {len(results)}\n")
            f.write(f"成功请求: {self.output_manager.stats['successful_requests']}\n")
            f.write(f"失败请求: {self.output_manager.stats['failed_requests']}\n")
            f.write("-" * 60 + "\n\n")

            # 按来源分组输出
            sources = {}
            for result in results:
                source = result['source'] if result['source'] else 'Unknown'
                if source not in sources:
                    sources[source] = []
                sources[source].append(result)

            for source, source_results in sources.items():
                f.write(f"📁 来源: {source}\n")
                f.write("-" * 30 + "\n")
                for result in source_results:
                    f.write(f"{result['url']}\n")
                f.write("\n")

    def _save_as_csv(self, results):
        """保存为CSV格式"""
        import csv
        with open(self.output_manager.output_file, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            # 写入头部
            writer.writerow(['URL', 'Source', 'Timestamp', 'Source_Type', 'Domain'])

            for result in results:
                url = result['url']
                source = result['source'] if result['source'] else 'Unknown'
                timestamp = result['timestamp']

                # 分析URL类型
                url_type = self._analyze_url_type(url)

                # 提取域名
                try:
                    domain = urlparse(url).netloc
                except:
                    domain = 'Unknown'

                writer.writerow([url, source, timestamp, url_type, domain])

    def _save_as_html(self, results, target_url):
        """保存为HTML格式"""
        html_content = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>API Finder - 扫描结果</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{ 
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; 
            margin: 20px; 
            background: linear-gradient(135deg, #f5f7fa 0%, #c3cfe2 100%);
            min-height: 100vh;
        }}
        .container {{ 
            max-width: 1200px; 
            margin: 0 auto; 
            background: white; 
            padding: 30px; 
            border-radius: 12px; 
            box-shadow: 0 8px 32px rgba(0,0,0,0.1);
        }}
        h1 {{ 
            color: #333; 
            text-align: center; 
            margin-bottom: 30px;
            font-size: 2.5em;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            background-clip: text;
        }}
        .stats {{ 
            display: grid; 
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); 
            gap: 20px; 
            margin: 30px 0; 
        }}
        .stat {{ 
            text-align: center; 
            padding: 20px; 
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            border-radius: 10px;
            color: white;
            transition: transform 0.3s ease;
        }}
        .stat:hover {{ transform: translateY(-5px); }}
        .stat-value {{ 
            font-size: 2em; 
            font-weight: bold; 
            margin-bottom: 5px;
        }}
        .stat-label {{ 
            font-size: 0.9em; 
            opacity: 0.9;
        }}
        .info-section {{
            margin: 30px 0;
            padding: 20px;
            background: #f8f9fa;
            border-radius: 8px;
            border-left: 4px solid #007bff;
        }}
        .info-section strong {{
            color: #007bff;
        }}
        table {{ 
            width: 100%; 
            border-collapse: collapse; 
            margin-top: 30px;
            background: white;
            border-radius: 8px;
            overflow: hidden;
            box-shadow: 0 4px 12px rgba(0,0,0,0.1);
        }}
        th, td {{ 
            padding: 15px; 
            text-align: left; 
            border-bottom: 1px solid #eee; 
        }}
        th {{ 
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            font-weight: 600;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }}
        tr:hover {{ 
            background-color: #f8f9ff; 
            transition: background-color 0.3s ease;
        }}
        .url-link {{ 
            color: #007bff; 
            text-decoration: none; 
            font-weight: 500;
            transition: all 0.3s ease;
            position: relative;
        }}
        .url-link:hover {{ 
            color: #0056b3;
            text-decoration: underline;
            transform: translateX(5px);
        }}
        .url-link::before {{
            content: '🔗';
            margin-right: 5px;
            opacity: 0.7;
        }}
        .source {{ 
            color: #666; 
            font-size: 0.9em;
            background: #e9ecef;
            padding: 4px 8px;
            border-radius: 4px;
            display: inline-block;
        }}
        .timestamp {{ 
            color: #888; 
            font-size: 0.85em;
            font-family: monospace;
        }}
        .filter-section {{
            margin: 30px 0;
            display: flex;
            gap: 15px;
            align-items: center;
            flex-wrap: wrap;
        }}
        .filter-section input {{ 
            padding: 12px 16px; 
            border: 2px solid #ddd; 
            border-radius: 25px; 
            font-size: 14px;
            transition: border-color 0.3s ease;
            flex: 1;
            min-width: 300px;
        }}
        .filter-section input:focus {{
            outline: none;
            border-color: #007bff;
            box-shadow: 0 0 0 3px rgba(0,123,255,0.1);
        }}
        .btn {{
            padding: 10px 20px;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            border: none;
            border-radius: 25px;
            cursor: pointer;
            font-size: 14px;
            transition: transform 0.3s ease;
        }}
        .btn:hover {{
            transform: translateY(-2px);
        }}
        .url-type {{
            font-size: 0.8em;
            padding: 3px 8px;
            border-radius: 12px;
            font-weight: 500;
            text-transform: uppercase;
        }}
        .url-type.api {{ background: #d4edda; color: #155724; }}
        .url-type.js {{ background: #fff3cd; color: #856404; }}
        .url-type.css {{ background: #d1ecf1; color: #0c5460; }}
        .url-type.image {{ background: #f8d7da; color: #721c24; }}
        .url-type.other {{ background: #e2e3e5; color: #383d41; }}
        .copy-btn {{
            background: transparent;
            border: 1px solid #007bff;
            color: #007bff;
            padding: 2px 8px;
            border-radius: 4px;
            cursor: pointer;
            font-size: 12px;
            margin-left: 10px;
            transition: all 0.3s ease;
        }}
        .copy-btn:hover {{
            background: #007bff;
            color: white;
        }}
        .footer {{
            margin-top: 50px;
            text-align: center;
            color: #666;
            font-size: 0.9em;
        }}
        @media (max-width: 768px) {{
            .stats {{ grid-template-columns: 1fr 1fr; }}
            .filter-section {{ flex-direction: column; }}
            .filter-section input {{ min-width: 100%; }}
        }}
    </style>
</head>
<body>
    <div class="container">
        <h1>🔍 API Finder 扫描结果</h1>

        <div class="stats">
            <div class="stat">
                <div class="stat-value">{len(results)}</div>
                <div class="stat-label">发现的URL</div>
            </div>
            <div class="stat">
                <div class="stat-value">{self.output_manager.stats['successful_requests']}</div>
                <div class="stat-label">成功请求</div>
            </div>
            <div class="stat">
                <div class="stat-value">{self.output_manager.stats['failed_requests']}</div>
                <div class="stat-label">失败请求</div>
            </div>
            <div class="stat">
                <div class="stat-value">{(datetime.now() - self.output_manager.stats['start_time']).total_seconds():.1f}s</div>
                <div class="stat-label">扫描用时</div>
            </div>
        </div>

        <div class="info-section">
            <p><strong>🎯 目标URL:</strong> <a href="{target_url}" target="_blank" class="url-link">{target_url}</a></p>
            <p><strong>🕐 扫描时间:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
            <p><strong>📊 扫描状态:</strong> {"✅ 完成" if len(results) > 0 else "⚠️ 未发现API端点"}</p>
        </div>

        <div class="filter-section">
            <input type="text" id="filterInput" placeholder="🔍 过滤URL（支持正则表达式）..." onkeyup="filterResults()">
            <button class="btn" onclick="exportResults()">📄 导出结果</button>
            <button class="btn" onclick="clearFilter()">🗑️ 清除过滤</button>
        </div>

        <table id="resultsTable">
            <thead>
                <tr>
                    <th>🔗 URL</th>
                    <th>📁 来源</th>
                    <th>🏷️ 类型</th>
                    <th>⏰ 时间</th>
                    <th>🛠️ 操作</th>
                </tr>
            </thead>
            <tbody>
"""

        for result in results:
            url = result['url']
            source = result['source'] if result['source'] else 'Unknown'
            timestamp = result['timestamp']
            url_type = self._analyze_url_type(url)

            # 格式化时间
            try:
                time_obj = datetime.fromisoformat(timestamp)
                formatted_time = time_obj.strftime('%H:%M:%S')
            except:
                formatted_time = timestamp

            # 生成类型标签的CSS类
            type_class = 'api' if 'api' in url.lower() else 'js' if '.js' in url else 'css' if '.css' in url else 'image' if any(
                ext in url.lower() for ext in ['.jpg', '.png', '.gif', '.svg']) else 'other'

            html_content += f"""
                <tr>
                    <td>
                        <a href="{url}" class="url-link" target="_blank" rel="noopener noreferrer">{url}</a>
                    </td>
                    <td><span class="source">{source.split('/')[-1] if source else 'Unknown'}</span></td>
                    <td><span class="url-type {type_class}">{url_type}</span></td>
                    <td><span class="timestamp">{formatted_time}</span></td>
                    <td>
                        <button class="copy-btn" onclick="copyToClipboard('{url}')">📋 复制</button>
                    </td>
                </tr>
"""

        html_content += f"""
            </tbody>
        </table>

        <div class="footer">
            <p>Generated by <strong>API Finder</strong> • {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
            <p>总共找到 <strong>{len(results)}</strong> 个URL</p>
        </div>
    </div>

    <script>
        function filterResults() {{
            const input = document.getElementById('filterInput');
            const filter = input.value.toLowerCase();
            const table = document.getElementById('resultsTable');
            const rows = table.getElementsByTagName('tr');

            for (let i = 1; i < rows.length; i++) {{
                const url = rows[i].getElementsByTagName('td')[0].textContent.toLowerCase();
                const source = rows[i].getElementsByTagName('td')[1].textContent.toLowerCase();
                const type = rows[i].getElementsByTagName('td')[2].textContent.toLowerCase();

                if (url.indexOf(filter) > -1 || source.indexOf(filter) > -1 || type.indexOf(filter) > -1) {{
                    rows[i].style.display = '';
                }} else {{
                    rows[i].style.display = 'none';
                }}
            }}
        }}

        function clearFilter() {{
            document.getElementById('filterInput').value = '';
            filterResults();
        }}

        function copyToClipboard(url) {{
            navigator.clipboard.writeText(url).then(function() {{
                // 创建提示
                const btn = event.target;
                const originalText = btn.textContent;
                btn.textContent = '✅ 已复制';
                btn.style.background = '#28a745';
                btn.style.color = 'white';

                setTimeout(() => {{
                    btn.textContent = originalText;
                    btn.style.background = 'transparent';
                    btn.style.color = '#007bff';
                }}, 1000);
            }});
        }}

        function exportResults() {{
            const table = document.getElementById('resultsTable');
            const rows = table.getElementsByTagName('tr');
            let csv = '序号,URL,来源,类型,时间\\n';

            for (let i = 1; i < rows.length; i++) {{
                if (rows[i].style.display !== 'none') {{
                    const cells = rows[i].getElementsByTagName('td');
                    const url = cells[0].textContent;
                    const source = cells[1].textContent;
                    const type = cells[2].textContent;
                    const time = cells[3].textContent;
                    csv += `${{i}},"${{url}}","${{source}}","${{type}}","${{time}}"\\n`;
                }}
            }}

            const blob = new Blob([csv], {{ type: 'text/csv;charset=utf-8;' }});
            const link = document.createElement('a');
            link.href = URL.createObjectURL(blob);
            link.download = 'api_finder_results.csv';
            link.click();
        }}

        // 添加键盘快捷键
        document.addEventListener('keydown', function(e) {{
            if (e.ctrlKey && e.key === 'f') {{
                e.preventDefault();
                document.getElementById('filterInput').focus();
            }}
        }});
    </script>
</body>
</html>
"""

        with open(self.output_manager.output_file, 'w', encoding='utf-8') as f:
            f.write(html_content)

    def _save_as_xml(self, results, target_url):
        """保存为XML格式"""
        xml_content = f"""<?xml version="1.0" encoding="UTF-8"?>
<api_scan_results>
    <metadata>
        <tool>Api-Finder</tool>
        <version>0.3.1</version>
        <scan_time>{datetime.now().isoformat()}</scan_time>
        <target_url>{target_url}</target_url>
        <total_results>{len(results)}</total_results>
    </metadata>
    <statistics>
        <successful_requests>{self.output_manager.stats['successful_requests']}</successful_requests>
        <failed_requests>{self.output_manager.stats['failed_requests']}</failed_requests>
        <api_endpoints>{self.output_manager.stats['api_endpoints']}</api_endpoints>
        <scan_duration>{(datetime.now() - self.output_manager.stats['start_time']).total_seconds():.1f}</scan_duration>
    </statistics>
    <results>
"""

        for result in results:
            url = result['url']
            source = result['source'] if result['source'] else 'Unknown'
            timestamp = result['timestamp']
            url_type = self._analyze_url_type(url)

            xml_content += f"""
        <result>
            <url><![CDATA[{url}]]></url>
            <source><![CDATA[{source}]]></source>
            <type>{url_type}</type>
            <timestamp>{timestamp}</timestamp>
        </result>
"""

        xml_content += """
    </results>
</api_scan_results>
"""

        with open(self.output_manager.output_file, 'w', encoding='utf-8') as f:
            f.write(xml_content)

    def _save_as_excel(self, results, target_url):
        """保存为Excel格式"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "API扫描结果"

            # 设置标题样式
            title_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")

            # 写入头部信息
            ws['A1'] = 'API Finder 扫描结果'
            ws['A1'].font = Font(bold=True, size=16)
            ws['A2'] = f'目标URL: {target_url}'
            ws['A3'] = f'扫描时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
            ws['A4'] = f'发现URL数量: {len(results)}'

            # 设置表格头部
            headers = ['URL', '来源', '类型', '域名', '时间戳']
            for i, header in enumerate(headers, 1):
                cell = ws.cell(row=6, column=i, value=header)
                cell.font = title_font
                cell.fill = header_fill

            # 写入数据
            for row, result in enumerate(results, 7):
                ws.cell(row=row, column=1, value=result['url'])
                ws.cell(row=row, column=2, value=result['source'] if result['source'] else 'Unknown')
                ws.cell(row=row, column=3, value=self._analyze_url_type(result['url']))
                try:
                    domain = urlparse(result['url']).netloc
                except:
                    domain = 'Unknown'
                ws.cell(row=row, column=4, value=domain)
                ws.cell(row=row, column=5, value=result['timestamp'])

            # 自动调整列宽
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(self.output_manager.output_file)

        except ImportError:
            self.output_manager.print_error("需要安装openpyxl库才能保存Excel格式: pip install openpyxl")
            # 回退到CSV格式
            self.output_manager.output_file = self.output_manager.output_file.rsplit('.', 1)[0] + '.csv'
            self._save_as_csv(results)

    def _save_as_markdown(self, results, target_url):
        """保存为Markdown格式"""
        md_content = f"""# 🔍 API Finder 扫描结果

## 📊 扫描信息

- **目标URL**: {target_url}
- **扫描时间**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
- **发现URL数量**: {len(results)}
- **成功请求**: {self.output_manager.stats['successful_requests']}
- **失败请求**: {self.output_manager.stats['failed_requests']}
- **扫描用时**: {(datetime.now() - self.output_manager.stats['start_time']).total_seconds():.1f}秒

## 🎯 发现的API端点

| URL | 来源 | 类型 | 时间 |
|-----|------|------|------|
"""

        for result in results:
            url = result['url']
            source = result['source'] if result['source'] else 'Unknown'
            source_display = source.split('/')[-1] if source else 'Unknown'
            url_type = self._analyze_url_type(url)

            try:
                time_obj = datetime.fromisoformat(result['timestamp'])
                formatted_time = time_obj.strftime('%H:%M:%S')
            except:
                formatted_time = result['timestamp']

            md_content += f"| {url} | {source_display} | {url_type} | {formatted_time} |\n"

        md_content += f"""

## 📈 统计信息

- 总URL数量: {len(results)}
- 唯一URL数量: {len(self._deduplicate_results())}
- 成功率: {round((self.output_manager.stats['successful_requests'] / max(1, self.output_manager.stats['successful_requests'] + self.output_manager.stats['failed_requests'])) * 100, 2)}%

---
*生成工具: Api-Finder v0.3.1*
"""

        with open(self.output_manager.output_file, 'w', encoding='utf-8') as f:
            f.write(md_content)

    def _analyze_url_type(self, url):
        """分析URL类型"""
        url_lower = url.lower()

        if any(keyword in url_lower for keyword in ['api', 'rest', 'graphql']):
            return 'API'
        elif url_lower.endswith('.json'):
            return 'JSON'
        elif url_lower.endswith('.xml'):
            return 'XML'
        elif any(keyword in url_lower for keyword in ['ajax', 'xhr']):
            return 'AJAX'
        elif any(keyword in url_lower for keyword in ['.php', '.jsp', '.asp']):
            return 'Dynamic'
        elif any(keyword in url_lower for keyword in ['.js', '.css', '.html']):
            return 'Static'
        else:
            return 'Other'